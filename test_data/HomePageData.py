import openpyxl
class HomePageData:

    test_HomePage_data = [{"firstname":"Abhishek", "email":"Gaikwad", "gender":"Male"},
                          {"firstname":"Akash", "email":"Verma", "gender":"Female"}]

    @staticmethod
    def getTestDataExcel(test_case_name):
        Dict = {}
        Book1 = openpyxl.load_workbook("F:\\SCODDEN LECTURE\\python notes\\udemy and youtube\\Book1.xlsx")
        sheet = Book1.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]






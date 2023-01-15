import openpyxl
Book1 = openpyxl.load_workbook("F:\\SCODDEN LECTURE\\python notes\\udemy and youtube\\Book1.xlsx")
sheet = Book1.active
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value="Abhishek" # to add value in excel file

print(sheet.cell(row=2, column=2).value) # to get the value

print(sheet.max_row) # to see max rows data in excel
print(sheet.max_column) # to see max column data in excel
print("------------------------------------------------")
# to get all values from 1 column and all rows till max rows
for i in range(1,sheet.max_row+1):
    print(sheet.cell(row = i, column = 1).value)
print("------------------------------------------------")
# to get all column and all rows data
for i in range(1,sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row = i, column = j).value)
print("------------------------------------------------")
# to get a selected row or column which matches the value testcase1
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i, column=1).value  == "testcase2":

        for j in range(2, sheet.max_column+1):

            Dict[sheet.cell(row = 1, column = j).value] = sheet.cell(row = i, column = j).value
print(Dict)






